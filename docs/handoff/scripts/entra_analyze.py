import json, unicodedata, re, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def norm(s):
    if not s: return ""
    s=unicodedata.normalize('NFKD',s).encode('ascii','ignore').decode()
    s=re.sub(r'[^A-Za-z0-9 ]',' ',s).upper()
    return re.sub(r'\s+',' ',s).strip()

def is_dni_mail(m):
    return bool(m) and bool(re.match(r'^[0-9]{8}@upeu\.edu\.pe$', m.lower()))
def is_real_upeu(m):
    return bool(m) and m.lower().endswith('@upeu.edu.pe') and not is_dni_mail(m)

users=json.load(open("/tmp/entra_users.json"))
# index by normalized display name
from collections import defaultdict
by_name=defaultdict(list)
for u in users:
    by_name[norm(u.get("display"))].append(u)

def parse_dt(s):
    if not s: return None
    try: return datetime.datetime.fromisoformat(s.replace("Z","+00:00"))
    except: return None

# ===== DUALES: personas con 2+ cuentas =====
duals=[]
for nm,us in by_name.items():
    if not nm: continue
    enabled=[u for u in us if u.get("enabled")]
    if len(enabled)>=2:
        # ordenar por último login interactivo desc (None al final)
        ranked=sorted(enabled, key=lambda u:(parse_dt(u.get("lastInteractive")) or datetime.datetime.min.replace(tzinfo=datetime.timezone.utc)), reverse=True)
        duals.append((nm,ranked))

# ===== CORREOS FALTANTES: match focos MidPoint sin correo -> Entra real =====
focos=[]
for line in open("/tmp/mp_sin_correo.txt"):
    line=line.rstrip("\n")
    if not line: continue
    p=line.split("|")
    if len(p)<6: continue
    focos.append({"codigo":p[0],"dni":p[1],"given":p[2],"surname":p[3],"email_mp":p[4],"aff":p[5]})

matches=[]
for f in focos:
    key=norm(f["given"]+" "+f["surname"])
    key2=norm(f["surname"]+" "+f["given"])
    cands=by_name.get(key,[]) or by_name.get(key2,[])
    # solo cuentas con correo real upeu y enabled, elegir la de login interactivo mas reciente
    real=[u for u in cands if u.get("enabled") and is_real_upeu(u.get("mail") or u.get("upn"))]
    if real:
        best=sorted(real, key=lambda u:(parse_dt(u.get("lastInteractive")) or datetime.datetime.min.replace(tzinfo=datetime.timezone.utc)), reverse=True)[0]
        matches.append((f, best, len(cands)))

# ===== EXCEL =====
NAVY="1F3864";BLUE="2E5496";GREEN="C6EFCE";AMBER="FFE699";RED="F8CBAD";WHITE="FFFFFF"
thin=Side(style='thin',color='BFBFBF');border=Border(left=thin,right=thin,top=thin,bottom=thin)
def hdr(ws,row,cols,fill=NAVY,h=26):
    for i,c in enumerate(cols,1):
        cell=ws.cell(row=row,column=i,value=c);cell.font=Font(bold=True,color="FFFFFF",size=10)
        cell.fill=PatternFill("solid",fgColor=fill);cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True);cell.border=border
    ws.row_dimensions[row].height=h
def put(ws,row,vals,fill=None):
    for i,v in enumerate(vals,1):
        cell=ws.cell(row=row,column=i,value=v);cell.alignment=Alignment(vertical='center',wrap_text=False,horizontal='left');cell.border=border
        if fill:cell.fill=PatternFill("solid",fgColor=fill)

wb=openpyxl.Workbook()

# Hoja duales
ws=wb.active;ws.title="Cuentas duplicadas"
ws.merge_cells('A1:H1');t=ws['A1'];t.value="Cuentas Entra duplicadas (misma persona, 2+ cuentas) — principal = login interactivo más reciente"
t.font=Font(bold=True,size=12,color="FFFFFF");t.fill=PatternFill("solid",fgColor=NAVY);t.alignment=Alignment(horizontal='center',vertical='center');ws.row_dimensions[1].height=28
hdr(ws,3,["Persona","UPN","mail","Habilitada","Último login interactivo","Creada","VEREDICTO","Cuentas"])
r=4
for nm,ranked in sorted(duals,key=lambda x:x[0]):
    for idx,u in enumerate(ranked):
        verdict="PRINCIPAL (conservar)" if idx==0 and u.get("lastInteractive") else ("DESCARTAR" if idx>0 else "revisar (sin login interactivo)")
        fill=GREEN if idx==0 and u.get("lastInteractive") else (RED if idx>0 else AMBER)
        put(ws,r,[nm if idx==0 else "", u.get("upn"), u.get("mail"), "Sí" if u.get("enabled") else "No",
                  (u.get("lastInteractive") or "(nunca)")[:19], (u.get("created") or "")[:10], verdict, len(ranked)],fill=fill)
        r+=1
for i,w in enumerate([30,34,30,10,22,12,24,9],1): ws.column_dimensions[get_column_letter(i)].width=w
ws.freeze_panes="A4"

# Hoja correos recuperables
ws2=wb.create_sheet("Correos recuperables de Entra")
ws2.merge_cells('A1:H1');t=ws2['A1'];t.value="Focos MidPoint sin correo real → buzón institucional encontrado en Entra (match por nombre)"
t.font=Font(bold=True,size=12,color="FFFFFF");t.fill=PatternFill("solid",fgColor=NAVY);t.alignment=Alignment(horizontal='center',vertical='center');ws2.row_dimensions[1].height=28
hdr(ws2,3,["Código MP","DNI","Nombre","Afiliación","Email en MidPoint (fantasma)","→ Correo real en Entra","Último login interactivo","#cuentas match"],fill=BLUE)
r=4
for f,best,ncand in sorted(matches,key=lambda x:x[0]["surname"]):
    mail=best.get("mail") or best.get("upn")
    fill=GREEN if best.get("lastInteractive") else AMBER
    put(ws2,r,[f["codigo"],f["dni"],(f["given"]+" "+f["surname"]).strip(),f["aff"],f["email_mp"] or "(vacío)",mail,(best.get("lastInteractive") or "(nunca)")[:19],ncand],fill=fill)
    r+=1
for i,w in enumerate([12,11,30,12,28,30,22,14],1): ws2.column_dimensions[get_column_letter(i)].width=w
ws2.freeze_panes="A4"

# Hoja resumen
ws3=wb.create_sheet("Resumen",0)
ws3.merge_cells('A1:C1');t=ws3['A1'];t.value="Barrido Entra ID UPeU — gobernanza de correo (AuditLog.Read.All)"
t.font=Font(bold=True,size=13,color="FFFFFF");t.fill=PatternFill("solid",fgColor=NAVY);t.alignment=Alignment(horizontal='center',vertical='center');ws3.row_dimensions[1].height=30
rows=[
 ("Total usuarios Entra analizados",len(users)),
 ("Personas con cuentas DUPLICADAS (2+)",len(duals)),
 ("  → cuentas a DESCARTAR (no principales)",sum(len(r)-1 for _,r in duals)),
 ("Focos MidPoint sin correo real",len(focos)),
 ("  → con correo recuperable de Entra",len(matches)),
 ("  → SIN match en Entra (reportar a DTI)",len(focos)-len(matches)),
]
rr=3
for k,v in rows:
    ws3.cell(row=rr,column=1,value=k).alignment=Alignment(vertical='center')
    c=ws3.cell(row=rr,column=2,value=v);c.font=Font(bold=True);c.alignment=Alignment(horizontal='center')
    ws3.cell(row=rr,column=1).border=border;ws3.cell(row=rr,column=2).border=border
    rr+=1
ws3.column_dimensions['A'].width=44;ws3.column_dimensions['B'].width=14
ws3.cell(row=rr+1,column=1,value="Método: 'principal' = cuenta con login INTERACTIVO más reciente (el no-interactivo es refresco de tokens, engaña). Generado "+datetime.date.today().isoformat()).font=Font(italic=True,size=9,color="606060")

out="/Users/alberto/Downloads/Entra_Gobernanza_Correo_UPeU_2026-06-04.xlsx"
wb.save(out)
print("EXCEL:",out)
print("duales:",len(duals),"| descartar:",sum(len(r)-1 for _,r in duals),"| correos recuperables:",len(matches),"de",len(focos))
