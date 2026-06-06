import json, unicodedata, re, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TODAY=datetime.datetime(2026,6,5,tzinfo=datetime.timezone.utc)
def norm(s):
    if not s: return ""
    s=unicodedata.normalize('NFKD',s).encode('ascii','ignore').decode()
    return re.sub(r'\s+',' ',re.sub(r'[^A-Za-z0-9 ]',' ',s).upper()).strip()
def parse_dt(s):
    if not s: return None
    try: return datetime.datetime.fromisoformat(s.replace("Z","+00:00"))
    except: return None
def days_since(s):
    d=parse_dt(s)
    return None if not d else (TODAY-d).days

def local(addr):
    return addr.split("@")[0] if addr and "@" in addr else ""
def is_personal_upeu(addr):
    if not addr or not addr.lower().endswith("@upeu.edu.pe"): return False
    lp=local(addr).lower()
    if re.match(r'^\d{8}$',lp): return False  # DNI fantasma
    # excluir funcionales evidentes
    func=['noreply','no-reply','biblioteca','gerencia','mesa','soporte','informes','info','admin','sistemas','secretaria','decanato','tesoreria','contabilidad','rrhh','marketing','crai','test','prueba']
    if any(f in lp for f in func) and '.' not in lp: return False
    return True
def proper_format(addr):
    # formato institucional canónico: nombre.apellido (con punto)
    return '.' in local(addr)

# enriquecer DNI por nombre normalizado
mp={}
for line in open("/tmp/mp_personas.txt"):
    p=line.rstrip("\n").split("|")
    if len(p)<5: continue
    code,dni,given,sur,aff=p[0],p[1],p[2],p[3],p[4]
    for k in (norm(given+" "+sur), norm(sur+" "+given)):
        if k and k not in mp: mp[k]={"code":code,"dni":dni,"aff":aff}

users=json.load(open("/tmp/entra_users.json"))
from collections import defaultdict
by_name=defaultdict(list)
for u in users:
    if u.get("enabled"): by_name[norm(u.get("display"))].append(u)

USED_DAYS=90
cases=[]
for nm,accs in by_name.items():
    if not nm: continue
    # cuentas con buzón personal @upeu
    pers=[a for a in accs if is_personal_upeu(a.get("mail") or a.get("upn"))]
    # dedup por upn
    seen=set(); pu=[]
    for a in pers:
        up=(a.get("upn") or "").lower()
        if up in seen: continue
        seen.add(up); pu.append(a)
    if len(pu)<2: continue  # solo casos con 2+ correos personales upeu
    # ordenar por último login interactivo (más reciente primero)
    for a in pu: a["_d"]=days_since(a.get("lastInteractive"))
    ranked=sorted(pu, key=lambda a:(a["_d"] if a["_d"] is not None else 99999))
    used=[a for a in ranked if a["_d"] is not None and a["_d"]<=USED_DAYS]
    # determinar keeper + acción
    if len(used)>=2:
        verdict="CONSOLIDAR (ambas en uso)"
        keeper=next((a for a in used if proper_format(a.get("upn") or a.get("mail"))), used[0])
        losers=[a for a in pu if a is not keeper]
        accion="Migrar buzón a %s (formato correcto). Las demás → ALIAS de esa cuenta. Conservar 1 sola." % (keeper.get("upn"))
    elif len(used)==1:
        verdict="BORRAR no usada"
        keeper=used[0]
        losers=[a for a in pu if a is not keeper]
        fmt="" if proper_format(keeper.get("upn") or keeper.get("mail")) else " (OJO: la usada no tiene formato nombre.apellido → renombrar)"
        accion="Conservar %s (única en uso). Mesa de ayuda: BORRAR %s (sin uso)."%(keeper.get("upn"), ", ".join(a.get("upn") for a in losers))+fmt
    else:
        verdict="REVISAR (ninguna en uso reciente)"
        keeper=next((a for a in ranked if proper_format(a.get("upn") or a.get("mail"))), ranked[0])
        losers=[a for a in pu if a is not keeper]
        accion="Ninguna con login interactivo <%dd. Revisar manualmente; tentativo conservar %s (formato correcto)."%(USED_DAYS, keeper.get("upn"))
    info=mp.get(nm,{})
    cases.append({"nombre":nm.title(),"dni":info.get("dni",""),"codigo":info.get("code",""),"aff":info.get("aff",""),
                  "verdict":verdict,"keeper":keeper,"losers":losers,"all":ranked,"accion":accion})

# ===== EXCEL =====
NAVY="1F3864";BLUE="2E5496";GREEN="C6EFCE";AMBER="FFE699";RED="F8CBAD";GREY="F2F2F2";WHITE="FFFFFF"
thin=Side(style='thin',color='BFBFBF');border=Border(left=thin,right=thin,top=thin,bottom=thin)
def hdr(ws,row,cols,fill=NAVY,h=30):
    for i,c in enumerate(cols,1):
        cell=ws.cell(row=row,column=i,value=c);cell.font=Font(bold=True,color="FFFFFF",size=10)
        cell.fill=PatternFill("solid",fgColor=fill);cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True);cell.border=border
    ws.row_dimensions[row].height=h
def cell(ws,r,c,v,fill=None,bold=False,wrap=True):
    x=ws.cell(row=r,column=c,value=v);x.alignment=Alignment(vertical='center',wrap_text=wrap,horizontal='left');x.border=border
    if fill:x.fill=PatternFill("solid",fgColor=fill)
    if bold:x.font=Font(bold=True)

wb=openpyxl.Workbook()

# Hoja resumen
ws0=wb.active; ws0.title="Resumen"
ws0.merge_cells('A1:C1');t=ws0['A1'];t.value="Cuentas Entra UPeU — un usuario con 2 correos personales (caso tipo Tito)"
t.font=Font(bold=True,size=13,color="FFFFFF");t.fill=PatternFill("solid",fgColor=NAVY);t.alignment=Alignment(horizontal='center',vertical='center');ws0.row_dimensions[1].height=30
nC=sum(1 for c in cases if c["verdict"].startswith("CONSOLIDAR"))
nB=sum(1 for c in cases if c["verdict"].startswith("BORRAR"))
nR=sum(1 for c in cases if c["verdict"].startswith("REVISAR"))
rows=[("Total casos (usuario con 2+ correos personales @upeu)",len(cases)),
      ("CONSOLIDAR — ambas en uso (migrar a 1 + la otra como alias)",nC),
      ("BORRAR — solo usa una (mesa de ayuda borra la otra)",nB),
      ("REVISAR — ninguna con login reciente (<90 días)",nR)]
rr=3
for k,v in rows:
    cell(ws0,rr,1,k); x=ws0.cell(row=rr,column=2,value=v);x.font=Font(bold=True);x.alignment=Alignment(horizontal='center');x.border=border
    rr+=1
ws0.column_dimensions['A'].width=58;ws0.column_dimensions['B'].width=12
ws0.cell(row=rr+1,column=1,value="Criterio: 'usa' = login INTERACTIVO en los últimos 90 días (el no-interactivo es refresco de tokens). 'formato correcto' = nombre.apellido@upeu.edu.pe. Cruce DNI/código con Oracle/MidPoint. Generado 2026-06-05.").font=Font(italic=True,size=9,color="606060")

# Hoja detalle
ws=wb.create_sheet("Casos 2 correos")
hdr(ws,1,["Persona","DNI","Código","Afil.","Correo 1 (UPN)","Login interac. 1","Correo 2 (UPN)","Login interac. 2","VEREDICTO","ACCIÓN"])
r=2
order={"CONSOLIDAR":0,"BORRAR":1,"REVISAR":2}
for c in sorted(cases,key=lambda x:(order.get(x["verdict"].split()[0],9), x["nombre"])):
    accs=c["all"]
    a1=accs[0]; a2=accs[1] if len(accs)>1 else {}
    fillv=GREEN if c["verdict"].startswith("CONSOLIDAR") else (RED if c["verdict"].startswith("BORRAR") else AMBER)
    def li(a):
        d=a.get("_d"); return (a.get("lastInteractive") or "(nunca)")[:10]+(" (%dd)"%d if d is not None else "")
    cell(ws,r,1,c["nombre"]);cell(ws,r,2,c["dni"]);cell(ws,r,3,c["codigo"]);cell(ws,r,4,c["aff"])
    cell(ws,r,5,a1.get("upn"));cell(ws,r,6,li(a1));cell(ws,r,7,a2.get("upn"));cell(ws,r,8,li(a2) if a2 else "")
    cell(ws,r,9,c["verdict"],fill=fillv,bold=True);cell(ws,r,10,c["accion"])
    r+=1
for i,w in enumerate([26,11,11,9,30,18,30,18,24,52],1): ws.column_dimensions[get_column_letter(i)].width=w
ws.freeze_panes="A2"

out="/Users/alberto/Downloads/Entra_Duales_Correo_Personal_UPeU_2026-06-05.xlsx"
wb.save(out)
print("EXCEL:",out)
print("Total casos 2 correos personales:",len(cases),"| CONSOLIDAR:",nC,"| BORRAR no-usada:",nB,"| REVISAR:",nR)
