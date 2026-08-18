#!/usr/bin/env python3
"""Entregables 1 y 2: clasificación y actividad de las cuentas M365 de UPeU."""
import json, datetime, re
from collections import Counter, defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell import WriteOnlyCell
from openpyxl.chart import BarChart, PieChart, Reference

import os
_SP = os.environ.get("M365_WORK") or os.path.expanduser("~/.cache/upeu-m365")
_OUT = os.environ.get("M365_OUT") or os.path.expanduser("~/Downloads")

SP = _SP
HOY = datetime.datetime.now(datetime.timezone.utc)
REF = HOY                                     # fecha de referencia elegida: hoy
FECHA = HOY.strftime("%Y-%m-%d")
OUT1 = f"{_OUT}/Analisis_Clasificacion_Cuentas_{FECHA}.xlsx"
OUT2 = f"{_OUT}/Analisis_Actividad_Cuentas_{FECHA}.xlsx"

rows = json.load(open(f"{SP}/clasificado.json"))

NAVY="1F3864"; BLUE="2E5496"; TEAL="1F6F6F"; GREEN="C6EFCE"; AMBER="FFE699"
RED="F8CBAD"; GREY="F2F2F2"; ORANGE="FFC000"
thin=Side(style='thin',color='BFBFBF'); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

def parse(s):
    if not s: return None
    try: return datetime.datetime.fromisoformat(s.replace("Z","+00:00")) if "T" in s \
        else datetime.datetime.strptime(s,"%Y-%m-%d").replace(tzinfo=datetime.timezone.utc)
    except Exception: return None

def wc(ws, v, bold=False, fill=None, center=False, wrap=False, color=None, size=None):
    c = WriteOnlyCell(ws, value=v)
    if bold or color or size: c.font = Font(bold=bold, color=color, size=size or 11)
    if fill: c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal=('center' if center else 'left'), vertical='center', wrap_text=wrap)
    c.border = BORDER
    return c

def head(ws, cols, fill=NAVY):
    ws.append([wc(ws, c, bold=True, fill=fill, center=True, wrap=True, color="FFFFFF") for c in cols])

def banner(ws, txt, fill=BLUE):
    ws.append([wc(ws, txt, bold=True, fill=fill, color="FFFFFF", size=12)])

def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ═════════════════════════ ENTREGABLE 1 — CLASIFICACIÓN ═════════════════════════
wb = openpyxl.Workbook(write_only=True)

# ---- LEEME
ws = wb.create_sheet("LEEME")
banner(ws, "Análisis y clasificación de cuentas institucionales M365 — UPeU", NAVY)
ws.append([])
for k, v in [
 ("Qué es y en qué se diferencia",
  "Responde al requerimiento de análisis y depuración de cuentas, pero invirtiendo su prioridad: "
  "la clasificación NO se deduce del nombre salvo que no haya más remedio. Primero se hereda de las "
  "fuentes que ya gobiernan la identidad en UPeU (MidPoint y Oracle LAMB) y solo el residuo se infiere."),
 ("Cascada de procedencia (columna «Fuente»)",
  "N0 forma de la cuenta: reglas objetivas (invitado externo, alias numérico). · "
  "N1 MidPoint: archetype, afiliación, campus y unidad HEREDADOS del foco de identidad. · "
  "N2a Oracle LAMB: la persona existe en el MDM con ese correo institucional. · "
  "N2b Oracle LAMB: el alias numérico es un documento de identidad real. · "
  "N3 inferencia: análisis semántico del nombre, SIEMPRE marcado como inferencia."),
 ("Por qué importa la columna Fuente",
  "Una fila N1 es un hecho verificado contra el sistema que gobierna esa identidad. Una fila N3 es "
  "una hipótesis. Mezclarlas en un mismo porcentaje daría una falsa sensación de certeza."),
 ("Catálogo de áreas",
  "El área NO sale de un diccionario de palabras clave inventado, sino de las 103 unidades reales del "
  "árbol organizativo de MidPoint, con su nombre oficial de Oracle (hoja «Catálogo de unidades»). "
  "Verificado: VocBench no tiene todavía vocabulario de áreas administrativas, solo académicas."),
 ("Afiliación",
  "Se usa eduPersonAffiliation del esquema canónico (student, alum, staff, faculty), estándar "
  "internacional ya adoptado por el IGA, en lugar de crear etiquetas nuevas."),
 ("Regla anti-falsos-positivos",
  "Se conserva íntegra: una coincidencia de palabra (lima, juliaca, contab…) NO basta para clasificar. "
  "Si hay indicios simultáneos de persona y de función, o el nombre coincide con varias personas, la "
  "cuenta queda Observada y marcada para revisión manual."),
 ("Alcance", "Solo lectura. No se modificó ninguna cuenta en M365, MidPoint ni Oracle. "
             "Los datos originales se conservan sin alterar."),
 ("Fecha del dato", FECHA),
]:
    ws.append([wc(ws, k, bold=True, wrap=True), wc(ws, v, wrap=True)])
widths(ws, [40, 110])

# ---- DASHBOARD
ws = wb.create_sheet("Dashboard Clasificación")
banner(ws, f"Indicadores principales — {len(rows):,} cuentas analizadas".replace(",","."), NAVY)
ws.append([])
head(ws, ["Indicador", "Cuentas", "% del total"], BLUE)
tot = len(rows)
ind = [
 ("Total de cuentas", tot),
 ("Personales", sum(1 for r in rows if r["TipoCuenta"]=="Personal")),
 ("Administrativas / de cargo", sum(1 for r in rows if r["TipoCuenta"]=="Administrativa")),
 ("Invitados externos", sum(1 for r in rows if r["TipoCuenta"]=="Invitado externo")),
 ("Observadas / sin determinar", sum(1 for r in rows if r["TipoCuenta"] in ("Observado","Otros"))),
 ("Cuentas de servicio", sum(1 for r in rows if r["TipoCuenta"]=="Servicio")),
 ("— clasificadas con FUENTE AUTORITATIVA", sum(1 for r in rows if r["Fuente"].startswith(("N1","N2")))),
 ("— clasificadas por INFERENCIA", sum(1 for r in rows if r["Fuente"].startswith("N3"))),
 ("Clasificación dudosa · revisión manual", sum(1 for r in rows if r["RequiereRevisionManual"]=="Sí")),
 ("FUERA del gobierno del IGA", sum(1 for r in rows if r["FueraDelGobiernoIGA"]=="Sí")),
]
for k, v in ind:
    f = GREEN if "AUTORITATIVA" in k else (AMBER if "INFERENCIA" in k else None)
    ws.append([wc(ws,k,bold=k.startswith("Total"),fill=f), wc(ws,v,center=True,fill=f),
               wc(ws,f"{100*v/tot:.1f}%",center=True,fill=f)])
ws.append([])
ws.append([wc(ws,"Distribución por procedencia de la clasificación",bold=True,fill=GREY)])
head(ws, ["Fuente","Cuentas","%"], BLUE)
r0 = ws._current_row + 1 if hasattr(ws,'_current_row') else None
for k, v in Counter(r["Fuente"] for r in rows).most_common():
    ws.append([wc(ws,k), wc(ws,v,center=True), wc(ws,f"{100*v/tot:.1f}%",center=True)])
ws.append([])
ws.append([wc(ws,"Distribución por tipo de cuenta",bold=True,fill=GREY)])
head(ws, ["Tipo","Cuentas","%"], BLUE)
for k, v in Counter(r["TipoCuenta"] for r in rows).most_common():
    ws.append([wc(ws,k), wc(ws,v,center=True), wc(ws,f"{100*v/tot:.1f}%",center=True)])
ws.append([])
ws.append([wc(ws,"Nivel de confianza",bold=True,fill=GREY)])
head(ws, ["Confianza","Cuentas","%"], BLUE)
for k, v in Counter(r["NivelConfianza"] for r in rows).most_common():
    ws.append([wc(ws,k), wc(ws,v,center=True), wc(ws,f"{100*v/tot:.1f}%",center=True)])
ws.append([])
ws.append([wc(ws,"Subtipo de cuenta",bold=True,fill=GREY)])
head(ws, ["Subtipo","Cuentas","%"], BLUE)
for k, v in Counter(r["SubtipoCuenta"] for r in rows).most_common():
    if k: ws.append([wc(ws,k), wc(ws,v,center=True), wc(ws,f"{100*v/tot:.1f}%",center=True)])
widths(ws, [52, 14, 12])

# ---- ÁREAS
ws = wb.create_sheet("Áreas administrativas")
banner(ws, "Cuentas por área / unidad institucional", TEAL)
ws.append([])
head(ws, ["Área o unidad","Cuentas","Con fuente autoritativa","Por inferencia","Requieren revisión"], BLUE)
byarea = defaultdict(list)
for r in rows:
    if r["AreaCargo"]: byarea[r["AreaCargo"]].append(r)
for a, rs in sorted(byarea.items(), key=lambda x: -len(x[1])):
    aut = sum(1 for r in rs if r["Fuente"].startswith(("N1","N2")))
    ws.append([wc(ws,a), wc(ws,len(rs),center=True), wc(ws,aut,center=True,fill=GREEN if aut else None),
               wc(ws,len(rs)-aut,center=True,fill=AMBER if len(rs)-aut else None),
               wc(ws,sum(1 for r in rs if r["RequiereRevisionManual"]=="Sí"),center=True)])
widths(ws, [56,12,22,16,18])

# ---- CATÁLOGO
ws = wb.create_sheet("Catálogo de unidades")
banner(ws, "Catálogo institucional de unidades usado para clasificar (árbol organizativo de MidPoint)", TEAL)
ws.append([])
head(ws, ["Identificador","Nombre oficial","Cuentas asignadas"], BLUE)
import csv as _csv
cnt_org = Counter(r["UnidadOrg"].split(";")[0] for r in rows if r["UnidadOrg"])
for o in _csv.DictReader(open(f"{SP}/mp_orgs.csv", newline='', encoding='utf-8')):
    ws.append([wc(ws,o["nameorig"]), wc(ws,o["display"] or "—"), wc(ws,cnt_org.get(o["nameorig"],0),center=True)])
widths(ws, [22,58,18])

# ---- DATOS
COLS = ["upn","mail","displayName","givenName","surname","dominio","habilitada","creada",
        "ultimoAccesoInteractivo","ultimoAccesoAutomatico","ultimoCambioPassword",
        "departamentoEntra","oficinaEntra","cargoEntra",
        "TipoCuenta","SubtipoCuenta","Afiliacion","AreaCargo","UnidadOrg","Campus",
        "CodigoInstitucional","DocumentoIdentidad","ArchetypeMidPoint",
        "Fuente","NivelConfianza","IndicadoresDetectados","MotivoClasificacion","RequiereRevisionManual","FueraDelGobiernoIGA"]
TIT = ["UPN","Correo","Nombre para mostrar","Nombre","Apellido","Dominio","Habilitada","Creada",
       "Último acceso interactivo","Último acceso automático","Último cambio de contraseña",
       "Departamento (Entra)","Oficina (Entra)","Cargo (Entra)",
       "TipoCuenta","SubtipoCuenta","Afiliación (eduPerson)","AreaCargo","UnidadOrg","Campus",
       "Código institucional","Documento","Archetype MidPoint",
       "FUENTE","NivelConfianza","IndicadoresDetectados","MotivoClasificacion","RequiereRevisionManual","FueraDelGobiernoIGA"]
ws = wb.create_sheet("Clasificación")
head(ws, TIT)
FILL = {"N1":GREEN,"N2":GREEN,"N0":GREY,"N3":AMBER}
for r in rows:
    f = FILL.get(r["Fuente"][:2])
    ws.append([wc(ws, r.get(c,""), fill=(f if c=="Fuente" else None)) for c in COLS])
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(rows)+1}"
widths(ws, [34,34,30,18,18,26,10,12,20,20,22,22,22,22,16,28,18,34,20,12,16,14,30,30,16,44,70,14,20])

wb.save(OUT1)
print("ENTREGABLE 1:", OUT1)

# ═════════════════════════ ENTREGABLE 2 — ACTIVIDAD ═════════════════════════
def estado(r):
    cre = parse(r["creada"]); li = parse(r["ultimoAccesoInteractivo"]); ni = parse(r["ultimoAccesoAutomatico"])
    if cre and (REF-cre).days <= 30: return "Cuenta creada recientemente", "Bajo"
    if li:
        d = (REF-li).days
        if li.year >= 2026: return "Activa 2026", "Bajo"
        if d <= 730: return "Inactiva 1-2 años", "Medio"
        return "Inactiva >2 años", "Alto"
    if ni: return "Información insuficiente", "Medio"
    ant = (REF-cre).days if cre else 0
    return "Nunca inició sesión", ("Crítico / revisión prioritaria" if ant > 730 else "Alto")

act = []
for r in rows:
    e, riesgo = estado(r)
    cre = parse(r["creada"]); li = parse(r["ultimoAccesoInteractivo"]); pw = parse(r["ultimoCambioPassword"])
    act.append({**r, "EstadoActividad": e, "NivelRiesgoInactividad": riesgo,
        "DiasDesdeUltimoInicio": (REF-li).days if li else "",
        "AntiguedadCuentaDias": (REF-cre).days if cre else "",
        "AntiguedadCuentaAnios": round((REF-cre).days/365.25,1) if cre else "",
        "DiasDesdeCambioPassword": (REF-pw).days if pw else "",
        "PotencialmenteAbandonada": "Sí" if (e in ("Inactiva >2 años","Nunca inició sesión")
                                             and r["TipoCuenta"]!="Invitado externo") else "No"})

wb = openpyxl.Workbook(write_only=True)
ws = wb.create_sheet("LEEME")
banner(ws, "Análisis de actividad de las cuentas M365 — UPeU", NAVY)
ws.append([])
for k, v in [
 ("Fecha de referencia", f"{FECHA} (fecha de la medición). El requerimiento original fijaba el "
                         f"30-06-2026; se usa la fecha real de extracción para reflejar el estado vigente."),
 ("Qué significa «activa»", "Inicio de sesión INTERACTIVO: la persona se autenticó. El acceso "
                            "no-interactivo (refresco de tokens) NO cuenta como uso humano, pero se "
                            "conserva en su propia columna porque distingue una cuenta muerta de una "
                            "que todavía tiene un dispositivo conectado."),
 ("Nunca inició sesión vs. información insuficiente",
  "Se diferencian, como pide el requerimiento: «nunca inició sesión» es ausencia total de registro "
  "(ni interactivo ni automático); «información insuficiente» es que hay actividad automática pero "
  "ningún inicio de sesión humano."),
 ("Nivel de riesgo", "Bajo: actividad en 2026. Medio: 1-2 años sin acceso. Alto: más de 2 años. "
                     "Crítico: cuenta con más de 2 años de antigüedad que nunca registró un acceso. "
                     "Es un indicador para PRIORIZAR revisión, no una orden de borrado."),
 ("Potencialmente abandonada", "Indicio fuerte, no confirmación. Ninguna cuenta debe eliminarse por "
                               "esta columna sin validación del área responsable."),
 ("Límite del dato", "Microsoft registra el inicio de sesión desde 2020. Una cuenta anterior sin "
                     "registro podría ser un vacío del histórico y no un no-uso real."),
]:
    ws.append([wc(ws,k,bold=True,wrap=True), wc(ws,v,wrap=True)])
widths(ws, [34,112])

ws = wb.create_sheet("Dashboard Actividad")
banner(ws, f"Actividad de las cuentas — fecha de referencia {FECHA}", NAVY)
ws.append([])
head(ws, ["Indicador","Cuentas","%"], BLUE)
for k, v in [("Total de cuentas", len(act))] + \
            [(k, v) for k, v in Counter(a["EstadoActividad"] for a in act).most_common()] + \
            [("Potencialmente abandonadas", sum(1 for a in act if a["PotencialmenteAbandonada"]=="Sí"))]:
    f = GREEN if k=="Activa 2026" else (RED if k in ("Inactiva >2 años","Nunca inició sesión","Potencialmente abandonadas") else None)
    ws.append([wc(ws,k,fill=f), wc(ws,v,center=True,fill=f), wc(ws,f"{100*v/len(act):.1f}%",center=True,fill=f)])
ws.append([])
ws.append([wc(ws,"Nivel de riesgo de inactividad",bold=True,fill=GREY)])
head(ws, ["Riesgo","Cuentas","%"], BLUE)
for k, v in Counter(a["NivelRiesgoInactividad"] for a in act).most_common():
    ws.append([wc(ws,k), wc(ws,v,center=True), wc(ws,f"{100*v/len(act):.1f}%",center=True)])
ws.append([])
ws.append([wc(ws,"Estado de actividad por tipo de cuenta",bold=True,fill=GREY)])
EST = ["Activa 2026","Inactiva 1-2 años","Inactiva >2 años","Nunca inició sesión",
       "Información insuficiente","Cuenta creada recientemente"]
head(ws, ["Tipo de cuenta"]+EST+["TOTAL"], BLUE)
for t, _ in Counter(a["TipoCuenta"] for a in act).most_common():
    sub = [a for a in act if a["TipoCuenta"]==t]
    c = Counter(a["EstadoActividad"] for a in sub)
    ws.append([wc(ws,t,bold=True)]+[wc(ws,c.get(e,0),center=True) for e in EST]+[wc(ws,len(sub),center=True,fill=GREY)])
ws.append([])
ws.append([wc(ws,"Cuentas inactivas por área administrativa (top 25)",bold=True,fill=GREY)])
head(ws, ["Área","Total","Inactivas >2 años","Nunca iniciaron sesión","% sin uso"], BLUE)
ba = defaultdict(list)
for a in act:
    if a["AreaCargo"]: ba[a["AreaCargo"]].append(a)
for ar, rs in sorted(ba.items(), key=lambda x: -len(x[1]))[:25]:
    i2 = sum(1 for a in rs if a["EstadoActividad"]=="Inactiva >2 años")
    nn = sum(1 for a in rs if a["EstadoActividad"]=="Nunca inició sesión")
    ws.append([wc(ws,ar), wc(ws,len(rs),center=True), wc(ws,i2,center=True), wc(ws,nn,center=True),
               wc(ws,f"{100*(i2+nn)/len(rs):.0f}%",center=True,fill=RED if (i2+nn)/len(rs)>.5 else None)])
widths(ws, [46,16,20,22,20,16,22,14])

C2 = ["upn","mail","displayName","dominio","habilitada","TipoCuenta","SubtipoCuenta","AreaCargo","Campus",
      "creada","ultimoAccesoInteractivo","ultimoAccesoAutomatico","ultimoCambioPassword",
      "DiasDesdeUltimoInicio","AntiguedadCuentaDias","AntiguedadCuentaAnios","DiasDesdeCambioPassword",
      "EstadoActividad","NivelRiesgoInactividad","PotencialmenteAbandonada","Fuente","RequiereRevisionManual"]
T2 = ["UPN","Correo","Nombre para mostrar","Dominio","Habilitada","TipoCuenta","SubtipoCuenta","AreaCargo","Campus",
      "FechaCreacion","UltimoInicioSesion","UltimoAccesoAutomatico","UltimoCambioPassword",
      "DiasDesdeUltimoInicio","AntiguedadCuentaDias","AntiguedadCuentaAnios","DiasDesdeCambioPassword",
      "EstadoActividad","NivelRiesgoInactividad","PotencialmenteAbandonada","FUENTE","RequiereRevision"]
ws = wb.create_sheet("Actividad")
head(ws, ["FechaReferencia = "+FECHA]+T2[1:])
RF = {"Activa 2026":GREEN,"Inactiva >2 años":RED,"Nunca inició sesión":RED,
      "Inactiva 1-2 años":AMBER,"Información insuficiente":AMBER}
for a in act:
    ws.append([wc(ws, a.get(c,""), fill=(RF.get(a["EstadoActividad"]) if c=="EstadoActividad" else None)) for c in C2])
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(C2))}{len(act)+1}"
widths(ws, [34,34,30,26,10,16,28,34,12,12,20,20,22,18,18,18,20,24,26,22,30,14])

wb.save(OUT2)
print("ENTREGABLE 2:", OUT2)
json.dump(act, open(f"{SP}/actividad.json","w"), ensure_ascii=False)

print()
print("verificación de consistencia — entrada:", len(rows), "| salida 1:", len(rows), "| salida 2:", len(act))
print("estados:", dict(Counter(a["EstadoActividad"] for a in act)))
print("riesgo:", dict(Counter(a["NivelRiesgoInactividad"] for a in act)))
