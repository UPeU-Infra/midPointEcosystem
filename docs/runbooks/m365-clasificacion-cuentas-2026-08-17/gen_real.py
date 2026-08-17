#!/usr/bin/env python3
"""Estado real del tenant M365 de UPeU y su evolución en 2026, por dominio.
Todo lo que aparece aquí está medido contra Microsoft Graph. Cada hoja declara
si es una foto exacta o una reconstrucción, y por qué."""
import json, datetime, os
from collections import Counter, defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference

SP  = "/private/tmp/claude-501/-Users-alberto-proyectos-productos-iga/ca1dd53b-d6bb-4df4-95a3-78d21daf87b8/scratchpad"
HOY = datetime.datetime.now(datetime.timezone.utc)
FECHA = HOY.strftime("%Y-%m-%d")
OUT = f"/Users/alberto/Downloads/M365_UPeU_Estado_Real_{FECHA}.xlsx"

users = json.load(open(f"{SP}/entra_users.json"))
audit = json.load(open(f"{SP}/audit.json")) if os.path.exists(f"{SP}/audit.json") else []

NAVY="1F3864"; BLUE="2E5496"; TEAL="1F6F6F"; GREEN="C6EFCE"; AMBER="FFE699"
RED="F8CBAD"; GREY="F2F2F2"; ORANGE="FFC000"
thin=Side(style='thin',color='BFBFBF'); border=Border(left=thin,right=thin,top=thin,bottom=thin)

def hdr(ws,row,cols,fill=NAVY,h=34):
    for i,c in enumerate(cols,1):
        x=ws.cell(row=row,column=i,value=c)
        x.font=Font(bold=True,color="FFFFFF",size=10)
        x.fill=PatternFill("solid",fgColor=fill)
        x.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        x.border=border
    ws.row_dimensions[row].height=h

def cell(ws,r,c,v,fill=None,bold=False,num=False):
    x=ws.cell(row=r,column=c,value=v)
    x.alignment=Alignment(vertical='center',horizontal=('center' if num else 'left'),wrap_text=not num)
    x.border=border
    if fill: x.fill=PatternFill("solid",fgColor=fill)
    if bold: x.font=Font(bold=True)
    return x

def titulo(ws,txt,ncols,color=BLUE,h=28):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncols)
    t=ws.cell(row=1,column=1,value=txt)
    t.font=Font(bold=True,size=12,color="FFFFFF"); t.fill=PatternFill("solid",fgColor=color)
    t.alignment=Alignment(horizontal='center',vertical='center'); ws.row_dimensions[1].height=h

def parse(s):
    if not s: return None
    try: return datetime.datetime.fromisoformat(s.replace("Z","+00:00"))
    except Exception: return None

def dom(u):
    a=u.get("upn") or ""
    return a.split("@")[-1].lower() if "@" in a else "(sin dominio)"

def totales(ws,r,c0,c1,fill=GREY):
    for i in range(c0,c1+1):
        cell(ws,r,i,sum(ws.cell(row=x,column=i).value or 0 for x in range(3,r)),bold=True,fill=fill,num=True)

DOMS  = [d for d,_ in Counter(dom(u) for u in users).most_common()]
MESES = [f"2026-{m:02d}" for m in range(1,HOY.month+1)]
ANIOS = sorted({parse(u["created"]).year for u in users if parse(u.get("created"))})

wb = openpyxl.Workbook()

# ============================================================ LEEME
ws=wb.active; ws.title="LEEME"
ws.merge_cells('A1:B1')
t=ws['A1']; t.value=f"Cuentas M365 / Entra ID de UPeU — estado real al {FECHA}"
t.font=Font(bold=True,size=14,color="FFFFFF"); t.fill=PatternFill("solid",fgColor=NAVY)
t.alignment=Alignment(horizontal='center',vertical='center'); ws.row_dimensions[1].height=34

notas=[
 ("Qué es exacto (foto de hoy)",
  "Las hojas «Foto real hoy», «Último acceso», «Nunca usadas» y «Altas por año» son un censo "
  "completo del tenant leído hoy vía Microsoft Graph: 75.341 cuentas, una por una. No hay muestreo "
  "ni estimación."),
 ("Qué es reconstruido (y por tanto un mínimo)",
  "La hoja «Stock acumulado 2026» proyecta hacia atrás la fecha de creación de las cuentas que "
  "existen HOY. Una cuenta creada en marzo y borrada en junio no aparece en ningún lado. Por eso "
  "el stock reconstruido de meses pasados es siempre MENOR O IGUAL al que realmente hubo."),
 ("Por qué no hay un histórico exacto",
  "Microsoft sí publica conteos diarios reales de los últimos 180 días, pero esa API exige el "
  "permiso Reports.Read.All y la aplicación MidPoint-IGA-UPeU NO lo tiene (devuelve 403 "
  "S2SUnauthorized). Si se le concede ese permiso, la serie diaria real de los últimos 6 meses "
  "queda disponible y esta limitación desaparece."),
 ("Movimiento con bajas incluidas",
  "Solo los registros de auditoría del directorio incluyen borrados y deshabilitaciones, y "
  "Microsoft los conserva 30 días. Es la hoja «Movimiento real 30d»: ahí sí está el ciclo de vida "
  "completo, pero solo del último mes."),
 ("Qué significa «último acceso»",
  "Inicio de sesión INTERACTIVO: la persona se autenticó. Se distingue del no-interactivo, que es "
  "refresco automático de tokens y aparece incluso en cuentas que nadie abre. Microsoft registra "
  "este dato desde 2020; por eso las cuentas creadas antes de esa fecha sin registro pueden ser "
  "un vacío del registro y no un no-uso real (son 76 cuentas, irrelevante para el total)."),
 ("Dominio de la cuenta",
  "Se toma del userPrincipalName, que es el identificador de inicio de sesión. Una cuenta puede "
  "recibir además correo en otros dominios como alias; eso se cuenta aparte."),
 ("Alcance", "Solo lectura. No se modificó ninguna cuenta ni ningún objeto del directorio."),
]
r=3
for k,v in notas:
    a=cell(ws,r,1,k,bold=True); a.alignment=Alignment(vertical='top',wrap_text=True)
    b=cell(ws,r,2,v); b.alignment=Alignment(vertical='top',wrap_text=True)
    ws.row_dimensions[r].height=max(30,15*(len(v)//94+1)); r+=1
ws.column_dimensions['A'].width=36; ws.column_dimensions['B'].width=104

# ============================================================ FOTO REAL HOY
ws=wb.create_sheet("Foto real hoy")
COLS=["Dominio (del UPN)","Total cuentas","Habilitadas","Deshabilitadas","Con buzón","Sin buzón",
      "Invitados externos","Con acceso en 2026","Sin usar NUNCA","Alias en otro dominio real"]
titulo(ws,f"Censo completo del tenant al {FECHA} — {len(users):,} cuentas en {len(DOMS)} dominios".replace(",","."),len(COLS))
hdr(ws,2,COLS,h=44)
r=3
for d in DOMS:
    us=[u for u in users if dom(u)==d]
    act=sum(1 for u in us if (parse(u.get("lastInteractive")) or datetime.datetime(1,1,1,tzinfo=datetime.timezone.utc)).year>=2026)
    nunca=sum(1 for u in us if not parse(u.get("lastInteractive")) and not parse(u.get("lastNonInteractive")))
    # alias reales: se excluye el dominio inicial onmicrosoft, que Microsoft añade
    # automáticamente a todo buzón y no significa multi-dominio
    alias=sum(1 for u in us for p in (u.get("proxy") or [])
              if "@" in p and p.split("@")[-1].lower() not in (d,"upeuedupe.onmicrosoft.com"))
    vals=[len(us),sum(1 for u in us if u["enabled"]),sum(1 for u in us if not u["enabled"]),
          sum(1 for u in us if u.get("mail")),sum(1 for u in us if not u.get("mail")),
          sum(1 for u in us if "#EXT#" in (u.get("upn") or "")),act,nunca,alias]
    cell(ws,r,1,d,bold=True)
    for i,v in enumerate(vals,2):
        cell(ws,r,i,v,num=True,fill=(RED if i==9 and v>1000 else None))
    r+=1
cell(ws,r,1,"TOTAL",bold=True,fill=GREY); totales(ws,r,2,len(COLS))
for i,w in enumerate([32,14,13,15,13,13,17,18,16,20],1): ws.column_dimensions[get_column_letter(i)].width=w
ws.freeze_panes="B3"

# ============================================================ ULTIMO ACCESO
ws=wb.create_sheet("Último acceso")
YEARS=[y for y in range(2020,HOY.year+1)]
COLS=["Dominio"]+[f"último acceso {y}" for y in YEARS]+["Sin registro","TOTAL"]
titulo(ws,"Año del último inicio de sesión INTERACTIVO — dato medido cuenta por cuenta",len(COLS),TEAL)
hdr(ws,2,COLS,h=40)
r=3
for d in DOMS:
    us=[u for u in users if dom(u)==d]
    c=Counter()
    for u in us:
        li=parse(u.get("lastInteractive"))
        c[li.year if li else "sin"]+=1
    cell(ws,r,1,d,bold=True)
    for i,y in enumerate(YEARS,2):
        v=c.get(y,0)
        cell(ws,r,i,v,num=True,fill=(GREEN if y==HOY.year and v else (RED if y<=2023 and v>1000 else None)))
    cell(ws,r,len(YEARS)+2,c.get("sin",0),num=True,fill=ORANGE if c.get("sin",0)>1000 else None)
    cell(ws,r,len(YEARS)+3,len(us),bold=True,num=True,fill=GREY)
    r+=1
cell(ws,r,1,"TOTAL",bold=True,fill=GREY); totales(ws,r,2,len(COLS))
r+=1
cell(ws,r,1,"% del tenant",bold=True)
for i in range(2,len(COLS)+1):
    v=ws.cell(row=r-1,column=i).value or 0
    cell(ws,r,i,f"{100*v/len(users):.1f}%",bold=True,num=True)
ws.column_dimensions['A'].width=32
for i in range(2,len(COLS)+1): ws.column_dimensions[get_column_letter(i)].width=17
ws.freeze_panes="B3"

# ============================================================ NUNCA USADAS
nunca=[u for u in users if not parse(u.get("lastInteractive")) and not parse(u.get("lastNonInteractive"))]
ws=wb.create_sheet("Nunca usadas")
COLS=["Año de creación","Cuentas nunca usadas","…habilitadas","…con buzón creado"]+DOMS
titulo(ws,f"{len(nunca):,} cuentas SIN NINGÚN inicio de sesión registrado — ni interactivo ni automático".replace(",","."),len(COLS),"C00000")
hdr(ws,2,COLS,h=40)
r=3
for y in ANIOS:
    sub=[u for u in nunca if (parse(u.get("created")) or datetime.datetime(1,1,1,tzinfo=datetime.timezone.utc)).year==y]
    if not sub: continue
    cell(ws,r,1,y,bold=True)
    cell(ws,r,2,len(sub),num=True,fill=(RED if len(sub)>1000 else None))
    cell(ws,r,3,sum(1 for u in sub if u["enabled"]),num=True)
    cell(ws,r,4,sum(1 for u in sub if u.get("mail")),num=True)
    cd=Counter(dom(u) for u in sub)
    for i,d in enumerate(DOMS,5): cell(ws,r,i,cd.get(d,0),num=True)
    r+=1
cell(ws,r,1,"TOTAL",bold=True,fill=GREY); totales(ws,r,2,len(COLS))
for i,w in enumerate([16,22,16,20]+[24]*len(DOMS),1): ws.column_dimensions[get_column_letter(i)].width=w
ws.freeze_panes="B3"

# ============================================================ INVITADOS EXTERNOS
inv=[u for u in users if "#EXT#" in (u.get("upn") or "")]
if inv:
    ws=wb.create_sheet("Invitados externos")
    COLS=["Dominio de origen del invitado","Invitados","Con acceso en 2026","Sin usar nunca","Creados en 2026"]
    titulo(ws,f"{len(inv)} invitados externos (colaboración B2B) — su UPN vive en el dominio inicial "
              f"onmicrosoft, pero su identidad real es de otra organización",len(COLS),TEAL)
    hdr(ws,2,COLS,h=40)
    orig=Counter((u["upn"] or "").split("#EXT#")[0].split("_")[-1].lower() for u in inv)
    r=3
    for d,_ in orig.most_common():
        sub=[u for u in inv if (u["upn"] or "").split("#EXT#")[0].split("_")[-1].lower()==d]
        act=sum(1 for u in sub if (parse(u.get("lastInteractive")) or datetime.datetime(1,1,1,tzinfo=datetime.timezone.utc)).year>=2026)
        nun=sum(1 for u in sub if not parse(u.get("lastInteractive")) and not parse(u.get("lastNonInteractive")))
        n26=sum(1 for u in sub if (parse(u.get("created")) or datetime.datetime(1,1,1,tzinfo=datetime.timezone.utc)).year==2026)
        cell(ws,r,1,d,bold=True)
        for i,v in enumerate([len(sub),act,nun,n26],2):
            cell(ws,r,i,v,num=True,fill=(RED if i==4 and v>100 else None))
        r+=1
    cell(ws,r,1,"TOTAL",bold=True,fill=GREY); totales(ws,r,2,len(COLS))
    for i,w in enumerate([36,14,20,18,18],1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="B3"

# ============================================================ ALTAS POR AÑO
ws=wb.create_sheet("Altas por año")
COLS=["Año de creación"]+DOMS+["TOTAL"]
titulo(ws,"Cuentas creadas cada año que siguen existiendo hoy — muestra de dónde viene el tenant actual",len(COLS))
hdr(ws,2,COLS,h=40)
r=3
for y in ANIOS:
    sub=[u for u in users if (parse(u.get("created")) or datetime.datetime(1,1,1,tzinfo=datetime.timezone.utc)).year==y]
    cell(ws,r,1,y,bold=True)
    cd=Counter(dom(u) for u in sub)
    for i,d in enumerate(DOMS,2):
        v=cd.get(d,0)
        cell(ws,r,i,v,num=True,fill=(AMBER if v>5000 else None))
    cell(ws,r,len(DOMS)+2,len(sub),bold=True,num=True,fill=GREY)
    r+=1
cell(ws,r,1,"TOTAL",bold=True,fill=GREY); totales(ws,r,2,len(COLS))
ws.column_dimensions['A'].width=16
for i in range(2,len(COLS)+1): ws.column_dimensions[get_column_letter(i)].width=24
ch=BarChart(); ch.type="col"; ch.title="Cuentas creadas por año (que siguen existiendo)"
ch.y_axis.title="cuentas"; ch.x_axis.title="año"; ch.height=9; ch.width=22
ch.add_data(Reference(ws,min_col=len(DOMS)+2,min_row=2,max_row=r-1),titles_from_data=True)
ch.set_categories(Reference(ws,min_col=1,min_row=3,max_row=r-1))
ws.add_chart(ch,f"A{r+3}")

# ============================================================ ALTAS POR MES 2026
altas=defaultdict(int); previo=Counter()
for u in users:
    d=dom(u); c=parse(u.get("created"))
    if not c or c.year<2026: previo[d]+=1
    elif c.year==2026: altas[(f"2026-{c.month:02d}",d)]+=1

ws=wb.create_sheet("Altas por mes 2026")
COLS=["Mes"]+DOMS+["TOTAL MES"]
titulo(ws,"Altas mes a mes durante 2026, por dominio",len(COLS))
hdr(ws,2,COLS,h=40)
r=3
for m in MESES:
    cell(ws,r,1,m,bold=True); tot=0
    for i,d in enumerate(DOMS,2):
        v=altas.get((m,d),0); tot+=v
        cell(ws,r,i,v,num=True,fill=(GREEN if v>500 else (AMBER if v>100 else None)))
    cell(ws,r,len(DOMS)+2,tot,bold=True,num=True,fill=GREY); r+=1
cell(ws,r,1,"TOTAL 2026",bold=True,fill=GREY); totales(ws,r,2,len(COLS))
ws.column_dimensions['A'].width=13
for i in range(2,len(COLS)+1): ws.column_dimensions[get_column_letter(i)].width=24
ch=BarChart(); ch.type="col"; ch.title="Altas mensuales 2026 (total tenant)"
ch.y_axis.title="cuentas creadas"; ch.x_axis.title="mes"; ch.height=9; ch.width=22
ch.add_data(Reference(ws,min_col=len(DOMS)+2,min_row=2,max_row=r-1),titles_from_data=True)
ch.set_categories(Reference(ws,min_col=1,min_row=3,max_row=r-1))
ws.add_chart(ch,f"A{r+3}")

# ============================================================ STOCK ACUMULADO
ws=wb.create_sheet("Stock acumulado 2026")
COLS=["Corte"]+DOMS+["TOTAL"]
titulo(ws,"RECONSTRUIDO — es un MÍNIMO: no incluye cuentas que existieron y ya fueron borradas",len(COLS),ORANGE)
hdr(ws,2,COLS,h=40)
acum={d:previo[d] for d in DOMS}
cell(ws,3,1,"31-dic-2025",bold=True)
for i,d in enumerate(DOMS,2): cell(ws,3,i,acum[d],num=True)
cell(ws,3,len(DOMS)+2,sum(acum.values()),bold=True,num=True,fill=GREY)
r=4
for m in MESES:
    for d in DOMS: acum[d]+=altas.get((m,d),0)
    cell(ws,r,1,"fin "+m,bold=True)
    for i,d in enumerate(DOMS,2): cell(ws,r,i,acum[d],num=True)
    cell(ws,r,len(DOMS)+2,sum(acum.values()),bold=True,num=True,fill=GREY); r+=1
ws.column_dimensions['A'].width=14
for i in range(2,len(COLS)+1): ws.column_dimensions[get_column_letter(i)].width=24
ch=LineChart(); ch.title="Stock reconstruido del tenant durante 2026"
ch.y_axis.title="cuentas"; ch.x_axis.title="corte"; ch.height=9; ch.width=22
ch.add_data(Reference(ws,min_col=len(DOMS)+2,min_row=2,max_row=r-1),titles_from_data=True)
ch.set_categories(Reference(ws,min_col=1,min_row=3,max_row=r-1))
ws.add_chart(ch,f"A{r+3}")

# ============================================================ MOVIMIENTO REAL 30d
if audit:
    ws=wb.create_sheet("Movimiento real 30d")
    COLS=["Día","Altas","Borrados","Deshabilitados","Rehabilitados","Neto"]
    titulo(ws,"Movimiento REAL con bajas incluidas — registros de auditoría (Microsoft solo conserva 30 días)",len(COLS),"C00000")
    hdr(ws,2,COLS,h=30)
    c=Counter((e['activityDateTime'][:10],e['activityDisplayName']) for e in audit)
    r=3
    for d in sorted({k[0] for k in c}):
        a=c.get((d,'Add user'),0); b=c.get((d,'Delete user'),0)
        di=c.get((d,'Disable account'),0); en=c.get((d,'Enable account'),0)
        cell(ws,r,1,d,bold=True)
        for i,v in enumerate([a,b,di,en,a-b],2):
            cell(ws,r,i,v,num=True,fill=(GREEN if i==2 and v>100 else (RED if i==3 and v>10 else None)))
        r+=1
    cell(ws,r,1,"TOTAL",bold=True,fill=GREY); totales(ws,r,2,6)
    r+=2
    cell(ws,r,1,"Quién ejecuta estas operaciones",bold=True); r+=1
    hdr(ws,r,["Origen","Operaciones","","","",""],fill=BLUE,h=20); r+=1
    for k,v in Counter((e.get('initiatedBy',{}).get('user',{}) or {}).get('userPrincipalName')
                       or (e.get('initiatedBy',{}).get('app',{}) or {}).get('displayName') or '?'
                       for e in audit).most_common(10):
        cell(ws,r,1,k); cell(ws,r,2,v,num=True); r+=1
    for i,w in enumerate([34,14,14,16,16,14],1): ws.column_dimensions[get_column_letter(i)].width=w

wb.save(OUT)
print("EXCEL:",OUT)
print("cuentas:",len(users),"| dominios:",len(DOMS))
print("nunca usadas:",len(nunca),"| de esas con buzon:",sum(1 for u in nunca if u.get("mail")))
print("acceso 2026:",sum(1 for u in users if (parse(u.get('lastInteractive')) or datetime.datetime(1,1,1,tzinfo=datetime.timezone.utc)).year>=2026))
