#!/usr/bin/env python3
"""Genera el Excel de revisión para DTI, con una hoja por campus."""
import json, datetime, sys
from collections import Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SP = "/private/tmp/claude-501/-Users-alberto-proyectos-productos-iga/4cb5ae8a-606c-44ee-b762-9f9a872a7da2/scratchpad"
FECHA = "2026-08-03"
OUT = f"/Users/alberto/Downloads/Entra_Duales_Correo_por_Campus_UPeU_{FECHA}.xlsx"

grupos = json.load(open(f"{SP}/grupos.json"))

NAVY="1F3864"; BLUE="2E5496"; TEAL="1F6F6F"; GREEN="C6EFCE"; AMBER="FFE699"
RED="F8CBAD"; GREY="F2F2F2"; INPUT="DDEBF7"; WHITE="FFFFFF"
thin=Side(style='thin', color='BFBFBF'); border=Border(left=thin,right=thin,top=thin,bottom=thin)

def hdr(ws,row,cols,fill=NAVY,h=34):
    for i,c in enumerate(cols,1):
        x=ws.cell(row=row,column=i,value=c)
        x.font=Font(bold=True,color="FFFFFF",size=10)
        x.fill=PatternFill("solid",fgColor=fill)
        x.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        x.border=border
    ws.row_dimensions[row].height=h

def cell(ws,r,c,v,fill=None,bold=False,wrap=True,size=None):
    x=ws.cell(row=r,column=c,value=v)
    x.alignment=Alignment(vertical='center',wrap_text=wrap,horizontal='left')
    x.border=border
    if fill: x.fill=PatternFill("solid",fgColor=fill)
    if bold or size: x.font=Font(bold=bold,size=size or 11)
    return x

def li_txt(a):
    d=a.get("d")
    if d is None:
        dn=a.get("dn")
        return "nunca (interactivo)" + (f" · no-interac. {a['lni'][:10]}" if a.get("lni") else "")
    return f"{a['li'][:10]} ({d}d)"

wb=openpyxl.Workbook()

# ==================== HOJA LÉEME ====================
ws=wb.active; ws.title="LEEME"
ws.merge_cells('A1:B1')
t=ws['A1']; t.value="Cuentas de correo duplicadas en Microsoft 365 / Entra ID — revisión DTI"
t.font=Font(bold=True,size=14,color="FFFFFF"); t.fill=PatternFill("solid",fgColor=NAVY)
t.alignment=Alignment(horizontal='center',vertical='center'); ws.row_dimensions[1].height=34

txt=[
 ("¿Qué contiene este archivo?",
  "Personas del tenant upeu.edu.pe que tienen DOS O MÁS buzones personales @upeu.edu.pe activos "
  "a la vez (por ejemplo albinoramos@ y albino.ramos@). Hay una hoja por campus para repartir la "
  "revisión, más una hoja de casos sin campus determinado."),
 ("¿Qué se espera de DTI?",
  "Revisar caso por caso y (a) corregirlo directamente en M365, y/o (b) llenar las columnas azules "
  "(DECISIÓN DTI en adelante) y devolver este mismo archivo. Las columnas azules son las únicas "
  "editables previstas; el resto es el dato medido."),
 ("¿Cómo se determinó el campus?",
  "Cascada de tres fuentes: 1) campus registrado en MidPoint (campusStudent / campusWorker, que "
  "vienen de Oracle LAMB); 2) para egresados, que no tienen campus en MidPoint, la columna SEDE de "
  "DAVID.VW_PERSONA_EGRESADO en Oracle LAMB; 3) atributos del propio Entra (officeLocation / city / "
  "department). La columna «Fuente campus» dice cuál se usó en cada fila. Quedan 25 personas sin "
  "campus atribuible (focos archivados, sin código ni DNI) en la hoja SIN CAMPUS."),
 ("⚠ Hoja «OJO buzon de area»",
  "PRIORIDAD. Casos donde el correo personal quedó agrupado con un BUZÓN DE ÁREA (remuneraciones@, "
  "soporte@, calidad.facihed@...) porque ese buzón tiene como nombre para mostrar el de la persona "
  "que lo administra. NO son correos duplicados: si se borra «la que no se usa» se puede dejar sin "
  "buzón a un área entera. Lo correcto es cambiarle el displayName al buzón funcional para que lleve "
  "el nombre del área, y tratar el correo personal aparte."),
 ("¿Qué significa «en uso»?",
  "Login INTERACTIVO (la persona escribió su contraseña) en los últimos 90 días. El login "
  "no-interactivo NO cuenta: es refresco automático de tokens y aparece incluso en cuentas que "
  "nadie abre."),
 ("VEREDICTO — CONSOLIDAR",
  "Las dos cuentas se usan de verdad. Hay que migrar el buzón a la del formato correcto "
  "(nombre.apellido@upeu.edu.pe) y dejar la otra como ALIAS, no borrarla en seco: la persona "
  "recibe correo en ambas."),
 ("VEREDICTO — BORRAR no usada",
  "Solo una cuenta tiene uso real; la otra lleva más de 90 días sin que nadie entre. Es el caso "
  "más seguro de limpiar."),
 ("VEREDICTO — BORRAR con tokens vivos",
  "Igual que el anterior, pero la cuenta a borrar registra actividad NO-interactiva reciente: hay "
  "un dispositivo o aplicación todavía conectado (correo en el móvil, Outlook, un servicio). Se "
  "puede borrar, pero avisando antes: al eliminarla ese dispositivo deja de recibir correo sin "
  "explicación. Van marcadas en naranja."),
 ("VEREDICTO — REVISAR",
  "Ninguna de las cuentas tiene login interactivo reciente. NO borrar sin contactar a la persona: "
  "puede ser alguien que no usa el correo institucional pero lo tiene asignado."),
 ("⚠ Hoja «NO FUSIONAR homónimos»",
  "Casos donde dos correos parecidos NO son de la misma persona sino de DOS PERSONAS DISTINTAS "
  "con el mismo nombre (verificado: resuelven a DNI diferentes en MidPoint). Si se borra una de "
  "esas cuentas se borra el buzón de una persona real. Esa hoja es informativa: no se toca nada."),
 ("Hoja «No personales»",
  "Grupos donde ninguna cuenta corresponde a una persona en MidPoint: buzones de área "
  "(admisionj@, activosfijosj@...) y pools de licencias de laboratorio (labuno.01@ … labuno.36@, "
  "hasta 53 cuentas bajo un mismo nombre). No son correos duplicados de nadie; requieren criterio "
  "de DTI sobre si el pool o el buzón siguen vigentes."),
 ("Alcance de la medición",
  "Solo lectura. No se modificó nada en M365, Entra ID, MidPoint, Koha ni Oracle. El conector de "
  "MidPoint hacia Entra ID está en modo solo-lectura."),
 ("Fecha del dato", f"{FECHA} — tenant leído completo vía Microsoft Graph el mismo día."),
]
r=3
for k,v in txt:
    a=cell(ws,r,1,k,bold=True); a.alignment=Alignment(vertical='top',wrap_text=True)
    b=cell(ws,r,2,v); b.alignment=Alignment(vertical='top',wrap_text=True)
    ws.row_dimensions[r].height=max(30, 15*(len(v)//95+1))
    r+=1
ws.column_dimensions['A'].width=34; ws.column_dimensions['B'].width=105

# ==================== COLUMNAS DE DETALLE ====================
COLS=["Persona","DNI","Código","Afiliación","Estado","Campus","Fuente campus",
      "Correo 1 (el más usado)","Último login interactivo 1",
      "Correo 2","Último login interactivo 2","Otros correos","N° cuentas",
      "VEREDICTO","ACCIÓN SUGERIDA",
      "DECISIÓN DTI","CORREO A CONSERVAR","CORREO(S) A BORRAR / ALIAS",
      "¿EJECUTADO EN M365?","FECHA","RESPONSABLE","OBSERVACIONES"]
WIDTH=[28,11,11,10,9,11,17, 30,20, 30,20, 26,7, 22,54, 18,28,28,16,12,18,34]
FIRST_INPUT=16   # columna P

def fill_sheet(ws, items, titulo, color):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(COLS))
    t=ws.cell(row=1,column=1,value=titulo)
    t.font=Font(bold=True,size=12,color="FFFFFF"); t.fill=PatternFill("solid",fgColor=color)
    t.alignment=Alignment(horizontal='center',vertical='center'); ws.row_dimensions[1].height=26
    hdr(ws,2,COLS)
    for i in range(FIRST_INPUT,len(COLS)+1):
        c=ws.cell(row=2,column=i); c.fill=PatternFill("solid",fgColor=BLUE)
    order={"CONSOLIDAR":0,"BORRAR":1,"REVISAR":2,"NO":3}
    r=3
    for g in sorted(items,key=lambda x:(order.get(x["veredicto"].split()[0],9), x["nombre"])):
        accs=g["cuentas"]
        a1=accs[0]; a2=accs[1] if len(accs)>1 else None
        otros=", ".join(a["upn"] for a in accs[2:]) if len(accs)>2 else ""
        v=g["veredicto"]
        if "tokens vivos" in v:      f="FFC000"   # naranja: borrable, pero hay algo conectado
        elif v.startswith("CONSOLIDAR"): f=GREEN
        elif v.startswith("BORRAR"):     f=RED
        elif v.startswith("NO"):         f=GREY
        else:                            f=AMBER
        vals=[g["nombre"],g["dni"],g["codigo"],g["aff"],g["lifecycle"],g["campus"],g["campus_src"],
              a1["upn"],li_txt(a1),
              (a2["upn"] if a2 else ""),(li_txt(a2) if a2 else ""),
              otros,len(accs),v,g["accion"]]
        for c,val in enumerate(vals,1):
            cell(ws,r,c,val,fill=(f if c==14 else None),bold=(c==14))
        for c in range(FIRST_INPUT,len(COLS)+1):
            cell(ws,r,c,"",fill=INPUT)
        r+=1
    if r>3:
        dv=DataValidation(type="list",formula1='"CONSOLIDAR,BORRAR,MANTENER AMBAS,ALIAS,PENDIENTE"',allow_blank=True)
        ws.add_data_validation(dv); dv.add(f"P3:P{r-1}")
        dv2=DataValidation(type="list",formula1='"SI,NO,PARCIAL"',allow_blank=True)
        ws.add_data_validation(dv2); dv2.add(f"S3:S{r-1}")
    for i,w in enumerate(WIDTH,1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="A3"
    ws.auto_filter.ref=f"A2:{get_column_letter(len(COLS))}{max(r-1,3)}"
    return r-3

personas   = [g for g in grupos if g["naturaleza"]=="PERSONA"]
homonimos  = [g for g in grupos if g["naturaleza"]=="HOMONIMOS"]
mixtos     = [g for g in grupos if g["naturaleza"]=="MIXTO_FUNCIONAL"]
funcional  = [g for g in grupos if g["naturaleza"] in ("FUNCIONAL","SIN_IDENTIDAD","POOL")]

campus_order=["LIMA","JULIACA","TARAPOTO"]
sheets={}
for camp in campus_order:
    items=[g for g in personas if g["campus"]==camp]
    if items:
        ws=wb.create_sheet(f"{camp} ({len(items)})")
        sheets[camp]=fill_sheet(ws,items,f"Campus {camp} — {len(items)} personas con 2+ correos personales",BLUE)

sin=[g for g in personas if g["campus"] not in campus_order]
if sin:
    ws=wb.create_sheet(f"SIN CAMPUS ({len(sin)})")
    sheets["SIN CAMPUS"]=fill_sheet(ws,sin,
        f"SIN CAMPUS DETERMINADO — {len(sin)} personas · investigar a qué sede pertenecen antes de accionar",TEAL)

if mixtos:
    ws=wb.create_sheet(f"OJO buzon de area ({len(mixtos)})")
    fill_sheet(ws,mixtos,
        "⚠ NO BORRAR SIN VERIFICAR — el correo personal está agrupado con un BUZÓN DE ÁREA que lleva "
        "el nombre de la persona como displayName. Borrar la cuenta «sin uso» puede dejar sin buzón a un área.",
        "C00000")

if homonimos:
    ws=wb.create_sheet(f"NO FUSIONAR ({len(homonimos)})")
    fill_sheet(ws,homonimos,
        "⚠ HOMÓNIMOS — personas DISTINTAS con el mismo nombre (DNI diferente en MidPoint). "
        "NO borrar ninguna cuenta. Hoja informativa.","C00000")

if funcional:
    ws=wb.create_sheet(f"No personales ({len(funcional)})")
    fill_sheet(ws,funcional,
        "Buzones de área / rol y pools de licencias — ninguna cuenta corresponde a una persona en "
        "MidPoint. No son correos duplicados de nadie. Criterio de DTI.","7F7F7F")

# ==================== RESUMEN ====================
ws=wb.create_sheet("Resumen",1)
ws.merge_cells('A1:E1')
t=ws['A1']; t.value=f"Resumen — cuentas duplicadas Entra ID UPeU · dato del {FECHA}"
t.font=Font(bold=True,size=13,color="FFFFFF"); t.fill=PatternFill("solid",fgColor=NAVY)
t.alignment=Alignment(horizontal='center',vertical='center'); ws.row_dimensions[1].height=30

hdr(ws,3,["Campus","CONSOLIDAR (ambas en uso)","BORRAR (una sin uso)","…de ellas con tokens vivos (avisar antes)","REVISAR (ninguna en uso)","TOTAL"],fill=BLUE,h=44)
r=4
def vk(v):
    return "CONSOLIDAR" if v.startswith("CONSOLIDAR") else ("BORRAR" if v.startswith("BORRAR") else "REVISAR")
def fila(label, items, fill=None):
    global r
    c=Counter(vk(g["veredicto"]) for g in items)
    tok=sum(1 for g in items if "tokens vivos" in g["veredicto"])
    cell(ws,r,1,label,bold=True,fill=fill)
    for i,v in enumerate([c.get("CONSOLIDAR",0), c.get("BORRAR",0), tok, c.get("REVISAR",0)],2):
        x=cell(ws,r,i,v,fill=fill); x.alignment=Alignment(horizontal='center')
    x=cell(ws,r,6,len(items),bold=True,fill=fill); x.alignment=Alignment(horizontal='center')
    r+=1
for camp in campus_order+["SIN CAMPUS"]:
    items=[g for g in personas if (g["campus"]==camp if camp!="SIN CAMPUS" else g["campus"] not in campus_order)]
    if items: fila(camp, items)
fila("TOTAL PERSONAS", personas, fill=GREY)
r+=2
cell(ws,r,1,"Casos apartados (no son duplicados de una persona)",bold=True); r+=1
cell(ws,r,1,"⚠ Buzón de área mezclado — NO borrar sin verificar",fill=RED); x=cell(ws,r,6,len(mixtos),bold=True,fill=RED); x.alignment=Alignment(horizontal='center'); r+=1
cell(ws,r,1,"⚠ Homónimos — personas distintas, NO tocar",fill=RED); x=cell(ws,r,6,len(homonimos),bold=True,fill=RED); x.alignment=Alignment(horizontal='center'); r+=1
cell(ws,r,1,"Buzones de área / rol y pools de licencias"); x=cell(ws,r,6,len(funcional),bold=True); x.alignment=Alignment(horizontal='center'); r+=1
cell(ws,r,1,"TOTAL GRUPOS DETECTADOS",bold=True,fill=GREY); x=cell(ws,r,6,len(grupos),bold=True,fill=GREY); x.alignment=Alignment(horizontal='center')
r+=2
ws.cell(row=r,column=1,value="«En uso» = login interactivo en los últimos 90 días. Campus por cascada MidPoint → Oracle LAMB (egresados) → Entra. "
        "Medición de solo lectura: no se modificó nada en ningún sistema.").font=Font(italic=True,size=9,color="606060")
for i,w in enumerate([46,22,20,26,22,12],1): ws.column_dimensions[get_column_letter(i)].width=w

wb.save(OUT)
print("EXCEL:",OUT)
print("Personas:",len(personas)," Homónimos:",len(homonimos)," Funcionales:",len(funcional)," TOTAL:",len(grupos))
for k,v in sheets.items(): print(f"  {k}: {v}")
